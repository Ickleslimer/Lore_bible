#!/usr/bin/env python3
"""One-time script: import Theriac Quests.xlsx → config/quest_song_seed.json.

Usage:
    python scripts/import_quest_songs.py --xlsx "Theriac Quests.xlsx" --out config/quest_song_seed.json

The spreadsheet has the following sheets:
  - Quest Master List (full list)
  - Per-character sheets (e.g., "Enoch Quests", "Izanami Quests", etc.)
Rows without a Quest ID are included; Quest ID is optional metadata.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore[assignment]


_IGNORE_SHEET_PATTERNS = re.compile(r"^sheet\d+$|^table\d+$", re.IGNORECASE)
BAND_HINTS: dict[str, str] = {
    # Enoch: Nine Inch Nails
    "every day is exactly the same": "nine_inch_nails",
    "into the void": "nine_inch_nails",
    "just like you imagined": "nine_inch_nails",
    "meet your master": "nine_inch_nails",
    "my violent heart": "nine_inch_nails",
    "ripe (with decay)": "nine_inch_nails",
    "ruiner": "nine_inch_nails",
    "sin": "nine_inch_nails",
    "the background world": "nine_inch_nails",
    "the beginning of the end": "nine_inch_nails",
    "the day the world went away": "nine_inch_nails",
    "the fragile": "nine_inch_nails",
    "the hand that feeds": "nine_inch_nails",
    "zero sum": "nine_inch_nails",
    # Izanami / Pandora: Guns N' Roses
    "14 years": "guns_n_roses",
    "breakdown": "guns_n_roses",
    "chinese democracy": "guns_n_roses",
    "paradise city": "guns_n_roses",
    "rocket queen": "guns_n_roses",
    "sweet child o' mine": "guns_n_roses",
    "sympathy for the devil": "guns_n_roses",
    "welcome to the jungle": "guns_n_roses",
    "eminence front": "the_who",
    "my generation": "the_who",
    "pinball wizard": "the_who",
    # Oyuun: Radiohead
    "exit music (for a film)": "radiohead",
    "fake plastic trees": "radiohead",
    "let down": "radiohead",
    "no surprises": "radiohead",
    "spectre": "radiohead",
    "the day i tried to live": "radiohead",
    # Beau: Meat Loaf / Bat Out Of Hell
    "bat out of hell": "meat_loaf",
    "brotherhood of man": "motley_crue",
    "love me forever": "motley_crue",
    # Ramasinta: Motörhead
    "killed by death": "motorhead",
    "motorhead": "motorhead",
    "stone dead forever": "motorhead",
    # RUINR: Black Sabbath / War Pigs
    "war pigs": "black_sabbath",
    "cochise": "audioslave",
    "show me how to live": "audioslave",
    # Altruism: Last Cup of Sorrow (Faith No More)
    "last cup of sorrow": "faith_no_more",
    "iliad": "odyssey_homer",
    "odyssey": "odyssey_homer",
    "man of war": "radiohead",
    "the good's gone": "the_who",
    "i'll be back": "other",
}


def _extract_band_hint(title: str) -> str:
    key = title.strip().lower()
    key = re.sub(r"\s+", " ", key)
    return BAND_HINTS.get(key, "other")


def _parse_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _characters_from_cell(value: object) -> list[str]:
    text = _parse_cell(value)
    if not text:
        return []
    parts = re.split(r"[,;/\n]+", text)
    out: list[str] = []
    for part in parts:
        name = part.strip().strip(".")
        if name and name.lower() not in ("", "none", "n/a", "tbd"):
            out.append(name)
    return out


def _main_character_from_cell(value: object) -> str:
    name = _parse_cell(value)
    if name.lower() in ("", "none", "n/a", "tbd", "protagonist"):
        return ""
    return name


def extract_quests_from_sheet(ws: object, seen_titles: set[str]) -> list[dict]:
    """Extract quest rows from an openpyxl worksheet."""
    try:
        rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
    except AttributeError:
        return []
    if not rows:
        return []

    header_row = rows[0]
    header_names = [str(h or "").strip().lower() for h in header_row]

    # Map expected columns
    col_map: dict[str, int] = {}
    for i, name in enumerate(header_names):
        if "quest title" in name or "quest" == name:
            col_map["title"] = i
        elif "main character" in name:
            col_map["main_char"] = i
        elif name in ("characters", "character"):
            col_map["characters"] = i
        elif "synopsis" in name or "description" in name:
            col_map["synopsis"] = i
        elif "quest id" in name or "id" == name:
            col_map["quest_id"] = i

    if "title" not in col_map:
        return []  # No title column found; skip sheet

    max_col = max(col_map.values())

    quests: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # Ensure row length accommodates the mapped column indices
        if len(row) <= max_col:
            continue
        title = _parse_cell(row[col_map["title"]])
        if not title:
            continue
        title_lower = title.lower()
        if title_lower in seen_titles:
            continue
        seen_titles.add(title_lower)

        entry: dict = {
            "quest_title": title,
            "main_character": "",
            "characters": [],
            "synopsis": "",
            "quest_id": "",
            "band_hint": _extract_band_hint(title),
        }
        if "main_char" in col_map and col_map["main_char"] < len(row):
            entry["main_character"] = _main_character_from_cell(row[col_map["main_char"]])
        if "characters" in col_map and col_map["characters"] < len(row):
            entry["characters"] = _characters_from_cell(row[col_map["characters"]])
        if "synopsis" in col_map and col_map["synopsis"] < len(row):
            entry["synopsis"] = _parse_cell(row[col_map["synopsis"]])
        if "quest_id" in col_map and col_map["quest_id"] < len(row):
            entry["quest_id"] = _parse_cell(row[col_map["quest_id"]])

        quests.append(entry)

    return quests


def infer_character(entry: dict, sheet_name: str) -> str:
    """Infer main_character from sheet name if not already populated."""
    if entry["main_character"]:
        return entry["main_character"]
    # Sheet names like "Izanami Quests" or "Enoch Quests" or "Altruism Quests"
    match = re.search(r"(\w+)\s+Quests", sheet_name, re.IGNORECASE)
    if match:
        candidate = match.group(1)
        if candidate.lower() not in ("quest", "theriac", "master", "all"):
            return candidate
    # Try extracting character name from sheet name directly
    cleaned = sheet_name.strip().lower()
    known_chars = ["izanami", "enoch", "oyuun", "pandora", "ramasinta", "beau", "joy", "altruism", "ruinr"]
    for char in known_chars:
        if char in cleaned:
            return char.capitalize()
    return ""


def run(xlsx_path: Path, out_path: Path) -> None:
    if load_workbook is None:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
        return

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    seen_titles: set[str] = set()
    all_quests: list[dict] = []
    sheet_names = wb.sheetnames if hasattr(wb, "sheetnames") else []

    for sheet_name in sheet_names:
        if _IGNORE_SHEET_PATTERNS.match(sheet_name):
            continue
        ws = wb[sheet_name]
        quests = extract_quests_from_sheet(ws, seen_titles)
        for entry in quests:
            if not entry["main_character"]:
                entry["main_character"] = infer_character(entry, sheet_name)
        all_quests.extend(quests)

    wb.close()

    # Deduplicate by quest_title, keeping the first occurrence (usually the master list)
    deduped: dict[str, dict] = {}
    for entry in all_quests:
        key = entry["quest_title"].strip().lower()
        if key not in deduped:
            deduped[key] = entry
        # If a later entry has more characters or synopsis, merge
        existing = deduped[key]
        if not existing["main_character"] and entry["main_character"]:
            existing["main_character"] = entry["main_character"]
        for char in entry.get("characters", []):
            if char and char not in existing["characters"]:
                existing["characters"].append(char)
        if not existing["synopsis"] and entry["synopsis"]:
            existing["synopsis"] = entry["synopsis"]
        if not existing["quest_id"] and entry["quest_id"]:
            existing["quest_id"] = entry["quest_id"]

    output = sorted(deduped.values(), key=lambda x: x["quest_title"].lower())

    payload = {"quest_song_seeds": output, "seed_count": len(output)}
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(output)} quest-song seeds to {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Theriac Quests spreadsheet -> quest_song_seed.json")
    parser.add_argument("--xlsx", type=Path, required=True, help="Path to the Theriac Quests.xlsx file")
    parser.add_argument("--out", type=Path, default=Path("config/quest_song_seed.json"), help="Output path for seed JSON")
    args = parser.parse_args()
    run(args.xlsx, args.out)


if __name__ == "__main__":
    main()